import openpyxl
import os
import sys
import json
import urllib.request
from datetime import datetime, time, date

sys.stdout.reconfigure(encoding='utf-8')

SUPABASE_URL = 'https://fmyafuhxlorbafbacywa.supabase.co'
SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZteWFmdWh4bG9yYmFmYmFjeXdhIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MjIyODU0NiwiZXhwIjoyMDc3ODA0NTQ2fQ.1wQC3L3gzGpZ2SwwQXMhXliZo_f7ye99vKEO7Q2iC5M'
EXCEL_PATH = os.path.expanduser('~') + "/Downloads/Invoice Monthly Update Report from Jan'26 updated.xlsx"

USER_MAP = {
    'sandeep': {'id': '04364421-22db-44e2-b1f1-08bcebc3d562', 'name': 'Sandeep B', 'role': 'operations'},
    'shilpa': {'id': '94a4f34e-f4d0-42d5-b2c5-7b43419a3325', 'name': 'Shilpa M', 'role': 'operations'},
    'issac': {'id': '2a637f75-8829-4572-bdbb-d60f7bb62bd6', 'name': 'Isaac Roy', 'role': 'operations'},
    'isaac': {'id': '2a637f75-8829-4572-bdbb-d60f7bb62bd6', 'name': 'Isaac Roy', 'role': 'operations'},
    'venkat': {'id': 'ed8ce87f-1f28-4a3a-a319-4dc502add40d', 'name': 'Venkatachalam', 'role': 'operations'},
    'harish': {'id': '10c74709-58d5-43d8-a8aa-9233699fc751', 'name': 'Harish H P', 'role': 'operations'},
    'murali': {'id': '63d47c04-28f5-43d4-8302-9ab990b7926a', 'name': 'Muralidhara B K', 'role': 'operations'},
    'muruli': {'id': '63d47c04-28f5-43d4-8302-9ab990b7926a', 'name': 'Muralidhara B K', 'role': 'operations'},
    'sashikanth': {'id': 'b6144e7e-d5e8-4f64-b50f-f15d01739c40', 'name': 'sashikanta das', 'role': 'operations'},
    'stany': {'id': '4f62cbdd-6410-4fa3-9d7e-04b05fc41866', 'name': 'Stany D Souza', 'role': 'operations'},
    'omkar': {'id': '8fb818f1-3750-4bd0-83ad-8832550c7fab', 'name': 'Omkar', 'role': 'operations'},
    'omkar/ stany': {'id': '8fb818f1-3750-4bd0-83ad-8832550c7fab', 'name': 'Omkar', 'role': 'operations'},
    'kannaiah': {'id': 'ffa83345-6b20-4995-af50-674790ecf906', 'name': 'Kannaiah R', 'role': 'operations'},
    'chandana': {'id': 'bbcbb70e-9c52-46c3-96e9-8e89155e35bd', 'name': 'Chandana R', 'role': 'hr'},
    'pooja': {'id': 'c96bc0e5-4b75-42a2-9f69-139de275ab7e', 'name': 'Poojashree S', 'role': 'hr'},
    'kavya': {'id': '07f61efd-24f2-457e-84b3-d8dafcb556c6', 'name': 'Kavya M', 'role': 'hr'},
    'arpitha': {'id': '6156dcf1-f6bb-4b5e-86d1-9236e6ec4a27', 'name': 'Arpitha Nairy', 'role': 'finance'},
    'sinchana': {'id': 'c9ffc969-8f2b-48cf-914c-0f30a661ba6f', 'name': 'Sinchana KM', 'role': 'finance'},
    'sudhan': {'id': '34aaddf9-77a0-4e2f-984b-b745cddb4808', 'name': 'Sudhan (IOT Architect)', 'role': 'finance_manager'}
}
DEFAULT_ADMIN = USER_MAP['sudhan']

SHEET_INFO = {
    "Jan'26": {'billing_month': '2026-01-01', 'created_at': '2026-02-15T12:00:00+00:00'},
    "Feb'26": {'billing_month': '2026-02-01', 'created_at': '2026-03-15T12:00:00+00:00'},
    "Mar'26": {'billing_month': '2026-03-01', 'created_at': '2026-04-15T12:00:00+00:00'},
    "April'26": {'billing_month': '2026-04-01', 'created_at': '2026-05-15T12:00:00+00:00'},
    "May'26": {'billing_month': '2026-05-01', 'created_at': '2026-06-15T12:00:00+00:00'},
    "June 2026": {'billing_month': '2026-06-01', 'created_at': '2026-07-15T12:00:00+00:00'}
}

def clean_str(val):
    if val is None: return ''
    return str(val).replace('\xa0', ' ').strip()

def parse_date_str(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.lower() == 'none': return None
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d-%b-%Y'):
        try:
            d = datetime.strptime(s, fmt)
            return d.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def parse_time_str(val):
    if val is None: return ''
    if isinstance(val, (time, datetime)):
        return val.strftime('%H:%M')
    s = str(val).strip()
    return '' if s.lower() == 'none' else s

def parse_num(val):
    if val is None: return 0.0
    try:
        s = str(val).replace(',', '').strip()
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def resolve_user(name_str):
    if not name_str: return DEFAULT_ADMIN
    return USER_MAP.get(str(name_str).strip().lower(), DEFAULT_ADMIN)

def normalize_title(name_str):
    if not name_str: return ''
    user = USER_MAP.get(str(name_str).strip().lower())
    if user:
        short_names = {
            '04364421-22db-44e2-b1f1-08bcebc3d562': 'Sandeep',
            '94a4f34e-f4d0-42d5-b2c5-7b43419a3325': 'Shilpa',
            '2a637f75-8829-4572-bdbb-d60f7bb62bd6': 'Isaac',
            'ed8ce87f-1f28-4a3a-a319-4dc502add40d': 'Venkat',
            '10c74709-58d5-43d8-a8aa-9233699fc751': 'Harish',
            '63d47c04-28f5-43d4-8302-9ab990b7926a': 'Murali',
            'b6144e7e-d5e8-4f64-b50f-f15d01739c40': 'Sashikanth',
            '4f62cbdd-6410-4fa3-9d7e-04b05fc41866': 'Stany',
            '8fb818f1-3750-4bd0-83ad-8832550c7fab': 'Omkar',
            'ffa83345-6b20-4995-af50-674790ecf906': 'Kannaiah',
            'bbcbb70e-9c52-46c3-96e9-8e89155e35bd': 'Chandana',
            'c96bc0e5-4b75-42a2-9f69-139de275ab7e': 'Pooja',
            '07f61efd-24f2-457e-84b3-d8dafcb556c6': 'Kavya',
            '6156dcf1-f6bb-4b5e-86d1-9236e6ec4a27': 'Arpitha',
            'c9ffc969-8f2b-48cf-914c-0f30a661ba6f': 'Sinchana',
        }
        return short_names.get(user['id'], user['name'])
    return str(name_str).strip()

def run_direct():
    print("Reading Excel and applying direct migration to Supabase...")
    
    # 1. Fetch existing keys from DB to skip duplicates
    headers = {
        'apikey': SERVICE_KEY,
        'Authorization': f'Bearer {SERVICE_KEY}',
        'Content-Type': 'application/json'
    }
    
    req1 = urllib.request.Request(f"{SUPABASE_URL}/rest/v1/site_invoice_tracker?select=site_name,manager_tentative_date&deleted_at=is.null&limit=2000", headers=headers)
    with urllib.request.urlopen(req1) as resp:
        existing_invoices = set(f"{r['site_name']}___{r.get('manager_tentative_date')}" for r in json.loads(resp.read().decode('utf-8')))
    print(f"Existing site_invoice_tracker records in DB: {len(existing_invoices)}")
    
    req2 = urllib.request.Request(f"{SUPABASE_URL}/rest/v1/site_finance_tracker?select=site_name,billing_month&deleted_at=is.null&limit=2000", headers=headers)
    with urllib.request.urlopen(req2) as resp:
        existing_finances = set(f"{r['site_name']}___{r.get('billing_month')}" for r in json.loads(resp.read().decode('utf-8')))
    print(f"Existing site_finance_tracker records in DB: {len(existing_finances)}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    new_invoices = []
    new_finances = []
    
    for sname in wb.sheetnames:
        if sname not in SHEET_INFO: continue
        ws = wb[sname]
        meta = SHEET_INFO[sname]
        b_month = meta['billing_month']
        c_at = meta['created_at']
        
        header_row_idx = None
        col_indices = {}
        for r_idx in range(1, 15):
            v1 = ws.cell(row=r_idx, column=1).value
            if v1 and str(v1).strip().lower() in ['sl.no', 'sl no', 'sl.no.']:
                header_row_idx = r_idx
                for c_idx in range(1, 45):
                    hv = ws.cell(row=r_idx, column=c_idx).value
                    if hv: col_indices[str(hv).strip()] = c_idx
                break
        if not header_row_idx: continue

        def get_col(*patterns):
            for p in patterns:
                for hn, ci in col_indices.items():
                    if p.lower() in hn.lower(): return ci
            return None

        col_site = get_col('SITES', 'ś', 'site')
        col_comp = get_col('Company Name', 'DESCRIPTION')
        col_cycle = get_col('Billing Cycle')
        col_ops = get_col('OPS INCHARGE')
        col_hr = get_col('HR INCHARGE')
        col_inv = get_col('INVOICE INCHARGE')
        col_mgr_tentative = 8
        col_mgr_received = 9
        col_hr_tentative = 11
        col_hr_received = 12
        col_att_time = 13
        col_inv_sharing = 15
        col_inv_prepared = 16
        col_inv_sent = 17
        col_inv_time = 19
        col_remarks = get_col('INVOICE SENT THROUGH Mail', 'Remarks')
        col_gross = get_col('GROSS')
        col_admin_chg = get_col('Admin Charges')
        col_cost_contract = get_col('As per Cost Sheet Taxable Value')
        col_cost_admin_fee = get_col('As per Cost Sheet  Admin Fees', 'As per Cost Sheet Admin Fees')
        col_admin_remarks = get_col('Admin fees Remarks')

        blank_streak = 0
        for r_idx in range(header_row_idx + 1, header_row_idx + 250):
            val_site = ws.cell(row=r_idx, column=col_site).value if col_site else None
            if not val_site or clean_str(val_site) == '':
                blank_streak += 1
                if blank_streak > 20: break
                continue
            site_name = clean_str(val_site)
            if site_name.lower() in ['sites', 'ś', 'description', 'sl.no']: continue
            blank_streak = 0
            
            comp_name = clean_str(ws.cell(row=r_idx, column=col_comp).value) if col_comp else 'PIFS'
            if not comp_name: comp_name = 'PIFS'
            cycle = clean_str(ws.cell(row=r_idx, column=col_cycle).value) if col_cycle else '3rd Billing Cycle'
            
            ops_val = clean_str(ws.cell(row=r_idx, column=col_ops).value) if col_ops else ''
            hr_val = clean_str(ws.cell(row=r_idx, column=col_hr).value) if col_hr else ''
            inv_val = clean_str(ws.cell(row=r_idx, column=col_inv).value) if col_inv else ''
            creator_user = resolve_user(inv_val)
            
            mgr_tentative = parse_date_str(ws.cell(row=r_idx, column=col_mgr_tentative).value)
            
            inv_key = f"{site_name}___{mgr_tentative}"
            if inv_key not in existing_invoices:
                new_invoices.append({
                    'site_name': site_name,
                    'company_name': comp_name,
                    'billing_cycle': cycle,
                    'ops_incharge': normalize_title(ops_val),
                    'hr_incharge': normalize_title(hr_val),
                    'invoice_incharge': normalize_title(inv_val),
                    'manager_tentative_date': mgr_tentative,
                    'manager_received_date': parse_date_str(ws.cell(row=r_idx, column=col_mgr_received).value),
                    'hr_tentative_date': parse_date_str(ws.cell(row=r_idx, column=col_hr_tentative).value),
                    'hr_received_date': parse_date_str(ws.cell(row=r_idx, column=col_hr_received).value),
                    'attendance_received_time': parse_time_str(ws.cell(row=r_idx, column=col_att_time).value),
                    'invoice_sharing_tentative_date': parse_date_str(ws.cell(row=r_idx, column=col_inv_sharing).value),
                    'invoice_prepared_date': parse_date_str(ws.cell(row=r_idx, column=col_inv_prepared).value),
                    'invoice_sent_date': parse_date_str(ws.cell(row=r_idx, column=col_inv_sent).value),
                    'invoice_sent_time': parse_time_str(ws.cell(row=r_idx, column=col_inv_time).value),
                    'invoice_sent_method_remarks': clean_str(ws.cell(row=r_idx, column=col_remarks).value) if col_remarks else '',
                    'created_by': creator_user['id'],
                    'created_by_name': creator_user['name'],
                    'created_by_role': creator_user['role'],
                    'created_at': c_at,
                    'updated_at': c_at
                })
                existing_invoices.add(inv_key)
                
            fin_key = f"{site_name}___{b_month}"
            if fin_key not in existing_finances:
                new_finances.append({
                    'site_name': site_name,
                    'company_name': comp_name,
                    'billing_month': b_month,
                    'contract_amount': parse_num(ws.cell(row=r_idx, column=col_cost_contract).value if col_cost_contract else 0),
                    'contract_management_fee': parse_num(ws.cell(row=r_idx, column=col_cost_admin_fee).value if col_cost_admin_fee else 0),
                    'billed_amount': parse_num(ws.cell(row=r_idx, column=col_gross).value if col_gross else 0),
                    'billed_management_fee': parse_num(ws.cell(row=r_idx, column=col_admin_chg).value if col_admin_chg else 0),
                    'remarks': clean_str(ws.cell(row=r_idx, column=col_admin_remarks).value) if col_admin_remarks else '',
                    'status': 'pending',
                    'created_by': creator_user['id'],
                    'created_by_name': creator_user['name'],
                    'created_by_role': creator_user['role'],
                    'created_at': c_at,
                    'updated_at': c_at
                })
                existing_finances.add(fin_key)
                
    wb.close()
    
    print(f"\nNew site_invoice_tracker records to insert: {len(new_invoices)}")
    print(f"New site_finance_tracker records to insert: {len(new_finances)}")
    
    # 2. Batch insert into Supabase via REST
    def batch_insert(table, records, batch_size=50):
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        for i in range(0, len(records), batch_size):
            chunk = records[i:i+batch_size]
            payload = json.dumps(chunk).encode('utf-8')
            req = urllib.request.Request(url, data=payload, headers=headers, method='POST')
            try:
                with urllib.request.urlopen(req) as resp:
                    pass
                print(f"  Inserted {table} rows {i+1} to {min(i+batch_size, len(records))}")
            except Exception as e:
                print(f"  Error inserting {table} batch: {e}")
                
    if new_invoices:
        print("\nInserting into site_invoice_tracker...")
        batch_insert('site_invoice_tracker', new_invoices)
        
    if new_finances:
        print("\nInserting into site_finance_tracker...")
        batch_insert('site_finance_tracker', new_finances)
        
    print("\n Direct migration completed successfully!")

if __name__ == '__main__':
    run_direct()
