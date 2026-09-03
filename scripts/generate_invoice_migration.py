import openpyxl
import os
import sys
import re
from datetime import datetime, time, date

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = os.path.expanduser('~') + "/Downloads/Invoice Monthly Update Report from Jan'26 updated.xlsx"
OUTPUT_SQL_PATH = "e:/backup/onboarding all files/Paradigm Office 4/supabase/migrations/20260902_migrate_invoice_records_jan_to_june.sql"

# Incharge user mapping
USER_MAP = {
    # Ops Incharge
    'sandeep': {'id': '04364421-22db-44e2-b1f1-08bcebc3d562', 'name': 'Sandeep B', 'role': 'operations'},
    'shilpa': {'id': '94a4f34e-f4d0-42d5-b2c5-7b43419a3325', 'name': 'Shilpa M', 'role': 'operations'},
    'issac': {'id': '2a637f75-8829-4572-bdbb-d60f7bb62bd6', 'name': 'Isaac Roy', 'role': 'operations'},
    'isaac': {'id': '2a637f75-8829-4572-bdbb-d60f7bb62bd6', 'name': 'Isaac Roy', 'role': 'operations'},
    'venkat': {'id': 'ed8ce87f-1f28-4a3a-a319-4dc502add40d', 'name': 'Venkatachalam', 'role': 'operations'},
    'harish': {'id': '10c74709-58d5-43d8-a8aa-9233699fc751', 'name': 'Harish H P', 'role': 'operations'},
    'murali': {'id': '63d47c04-28f5-43d4-8302-9ab990b7926a', 'name': 'Muralidhara B K', 'role': 'operations'},
    'muruli': {'id': '63d47c04-28f5-43d4-8302-9ab990b7926a', 'name': 'Muralidhara B K', 'role': 'operations'},
    'sashikanth': {'id': 'b6144e7e-d5e8-4f64-b50f-f15d01739c40', 'name': 'sashikanta das', 'role': 'operations'},
    'stany': {'id': '4f62cbdd-6410-4fa3-9d7e-04b05fc41866', 'name': 'Stany D Souza', 'role': 'operations'},
    'omkar': {'id': '8fb818f1-3750-4bd0-83ad-8832550c7fab', 'name': 'Omkar', 'role': 'operations'},
    'omkar/ stany': {'id': '8fb818f1-3750-4bd0-83ad-8832550c7fab', 'name': 'Omkar', 'role': 'operations'},
    'kannaiah': {'id': 'ffa83345-6b20-4995-af50-674790ecf906', 'name': 'Kannaiah R', 'role': 'operations'},

    # HR Incharge
    'chandana': {'id': 'bbcbb70e-9c52-46c3-96e9-8e89155e35bd', 'name': 'Chandana R', 'role': 'hr'},
    'pooja': {'id': 'c96bc0e5-4b75-42a2-9f69-139de275ab7e', 'name': 'Poojashree S', 'role': 'hr'},
    'kavya': {'id': '07f61efd-24f2-457e-84b3-d8dafcb556c6', 'name': 'Kavya M', 'role': 'hr'},

    # Invoice Incharge
    'arpitha': {'id': '6156dcf1-f6bb-4b5e-86d1-9236e6ec4a27', 'name': 'Arpitha Nairy', 'role': 'finance'},
    'sinchana': {'id': 'c9ffc969-8f2b-48cf-914c-0f30a661ba6f', 'name': 'Sinchana KM', 'role': 'finance'},
    'arya': {'id': 'de592a75-44de-441a-80fe-dc4a77e6901c', 'name': 'Arya Thomas', 'role': 'finance'},
    'sudhan': {'id': '34aaddf9-77a0-4e2f-984b-b745cddb4808', 'name': 'Sudhan (IOT Architect)', 'role': 'finance_manager'}
}

DEFAULT_ADMIN = USER_MAP['sudhan']

SHEET_INFO = {
    "Jan'26": {'billing_month': '2026-01-01', 'created_at': '2026-02-15 12:00:00+00'},
    "Feb'26": {'billing_month': '2026-02-01', 'created_at': '2026-03-15 12:00:00+00'},
    "Mar'26": {'billing_month': '2026-03-01', 'created_at': '2026-04-15 12:00:00+00'},
    "April'26": {'billing_month': '2026-04-01', 'created_at': '2026-05-15 12:00:00+00'},
    "May'26": {'billing_month': '2026-05-01', 'created_at': '2026-06-15 12:00:00+00'},
    "June 2026": {'billing_month': '2026-06-01', 'created_at': '2026-07-15 12:00:00+00'}
}

def clean_str(val):
    if val is None:
        return ''
    s = str(val).replace('\xa0', ' ').strip()
    return s

def sql_str(val):
    if val is None or str(val).strip() == '':
        return 'NULL'
    s = str(val).replace("'", "''").strip()
    return f"'{s}'"

def sql_date(val):
    if val is None:
        return 'NULL'
    if isinstance(val, (datetime, date)):
        return f"'{val.strftime('%Y-%m-%d')}'"
    s = str(val).strip()
    if not s or s.lower() == 'none':
        return 'NULL'
    # Try parsing string
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d-%b-%Y'):
        try:
            d = datetime.strptime(s, fmt)
            return f"'{d.strftime('%Y-%m-%d')}'"
        except ValueError:
            pass
    return 'NULL'

def sql_time(val):
    if val is None:
        return 'NULL'
    if isinstance(val, time):
        return f"'{val.strftime('%H:%M')}'"
    if isinstance(val, datetime):
        return f"'{val.strftime('%H:%M')}'"
    s = str(val).strip()
    if not s or s.lower() == 'none':
        return 'NULL'
    return sql_str(s)

def sql_num(val):
    if val is None:
        return '0'
    try:
        # remove commas
        s = str(val).replace(',', '').strip()
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return f"{f:.2f}"
    except (ValueError, TypeError):
        return '0'

def resolve_user(name_str):
    if not name_str:
        return DEFAULT_ADMIN
    key = str(name_str).strip().lower()
    return USER_MAP.get(key, DEFAULT_ADMIN)

def normalize_title(name_str):
    if not name_str:
        return ''
    key = str(name_str).strip().lower()
    user = USER_MAP.get(key)
    if user:
        # Return common short display name
        short_names = {
            '04364421-22db-44e2-b1f1-08bcebc3d562': 'Sandeep',
            '94a4f34e-f4d0-42d5-b2c5-7b43419a3325': 'Shilpa',
            '2a637f75-8829-4572-bdbb-d60f7bb62bd6': 'Isaac',
            'ed8ce87f-1f28-4a3a-a319-4dc502add40d': 'Venkat',
            '10c74709-58d5-43d8-a8aa-9233699fc751': 'Harish',
            '63d47c04-28f5-43d4-8302-9ab990b7926a': 'Murali',
            'b6144e7e-d5e8-4f64-b50f-f15d01739c40': 'Sashikanth',
            '4f62cbdd-6410-4fa3-9d7e-04b05fc41866': 'Stany',
            '8fb818f1-3750-4bd0-83ad-8832550c7fab': 'Omkar',
            'ffa83345-6b20-4995-af50-674790ecf906': 'Kannaiah',
            'bbcbb70e-9c52-46c3-96e9-8e89155e35bd': 'Chandana',
            'c96bc0e5-4b75-42a2-9f69-139de275ab7e': 'Pooja',
            '07f61efd-24f2-457e-84b3-d8dafcb556c6': 'Kavya',
            '6156dcf1-f6bb-4b5e-86d1-9236e6ec4a27': 'Arpitha',
            'c9ffc969-8f2b-48cf-914c-0f30a661ba6f': 'Sinchana',
        }
        return short_names.get(user['id'], user['name'])
    return str(name_str).strip()

def process():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    invoice_records = []
    finance_records = []
    
    for sname in wb.sheetnames:
        if sname not in SHEET_INFO:
            continue
        ws = wb[sname]
        meta = SHEET_INFO[sname]
        billing_month = meta['billing_month']
        created_at_val = meta['created_at']
        
        # 1. Find Header Row
        header_row_idx = None
        col_indices = {}
        for r_idx in range(1, 15):
            val1 = ws.cell(row=r_idx, column=1).value
            if val1 and str(val1).strip().lower() in ['sl.no', 'sl no', 'sl.no.']:
                header_row_idx = r_idx
                # Map headers
                for c_idx in range(1, 45):
                    h_val = ws.cell(row=r_idx, column=c_idx).value
                    if h_val:
                        h_clean = str(h_val).strip()
                        col_indices[h_clean] = c_idx
                break
                
        if not header_row_idx:
            print(f"Warning: Could not find header row in sheet {sname}")
            continue

        # Helper to get column index with fallbacks
        def get_col(*patterns):
            for p in patterns:
                for h_name, c_idx in col_indices.items():
                    if p.lower() in h_name.lower():
                        return c_idx
            return None

        col_site = get_col('SITES', 'ś', 'site')
        col_comp = get_col('Company Name', 'DESCRIPTION')
        col_cycle = get_col('Billing Cycle')
        col_ops = get_col('OPS INCHARGE')
        col_hr = get_col('HR INCHARGE')
        col_inv = get_col('INVOICE INCHARGE')
        
        # Dates & Status
        col_mgr_tentative = 8 # Tentative date for attendance submission (Manager)
        col_mgr_received = 9  # Attendance Received from Site Manager
        col_hr_tentative = 11 # Tentative date for attendance submission (HR)
        col_hr_received = 12  # Attendance Received from HR
        col_att_time = 13     # ATTENDANCE RECEIVED TIMINGS FROM HR
        col_inv_sharing_tentative = 15 # Tentative date for Invoice Raising
        col_inv_prepared = 16 # INVOICE PREPARED DATE
        col_inv_sent = 17     # INVOICE SENT DATE
        col_inv_time = 19     # Invoice Sent Timings
        
        col_remarks = get_col('INVOICE SENT THROUGH Mail', 'Remarks')
        
        # Finance fields
        col_gross = get_col('GROSS')
        col_admin_chg = get_col('Admin Charges')
        col_cost_contract = get_col('As per Cost Sheet Taxable Value')
        col_cost_admin_fee = get_col('As per Cost Sheet  Admin Fees', 'As per Cost Sheet Admin Fees')
        col_admin_remarks = get_col('Admin fees Remarks')

        # Read data rows
        blank_streak = 0
        for r_idx in range(header_row_idx + 1, header_row_idx + 250):
            val_sl = ws.cell(row=r_idx, column=1).value
            val_site = ws.cell(row=r_idx, column=col_site).value if col_site else None
            
            if val_site is None or clean_str(val_site) == '':
                blank_streak += 1
                if blank_streak > 20:
                    break
                continue
            
            site_name = clean_str(val_site)
            # Skip if it repeats header
            if site_name.lower() in ['sites', 'ś', 'description', 'sl.no']:
                continue
                
            blank_streak = 0
            
            comp_name = clean_str(ws.cell(row=r_idx, column=col_comp).value) if col_comp else 'PIFS'
            if not comp_name:
                comp_name = 'PIFS'
            cycle = clean_str(ws.cell(row=r_idx, column=col_cycle).value) if col_cycle else '3rd Billing Cycle'
            
            ops_val = clean_str(ws.cell(row=r_idx, column=col_ops).value) if col_ops else ''
            hr_val = clean_str(ws.cell(row=r_idx, column=col_hr).value) if col_hr else ''
            inv_val = clean_str(ws.cell(row=r_idx, column=col_inv).value) if col_inv else ''
            
            ops_incharge = normalize_title(ops_val)
            hr_incharge = normalize_title(hr_val)
            invoice_incharge = normalize_title(inv_val)
            
            # Audit user attribution: assign to Invoice Incharge
            creator_user = resolve_user(inv_val)
            
            mgr_tentative = ws.cell(row=r_idx, column=col_mgr_tentative).value
            mgr_received = ws.cell(row=r_idx, column=col_mgr_received).value
            hr_tentative = ws.cell(row=r_idx, column=col_hr_tentative).value
            hr_received = ws.cell(row=r_idx, column=col_hr_received).value
            att_time = ws.cell(row=r_idx, column=col_att_time).value
            inv_sharing = ws.cell(row=r_idx, column=col_inv_sharing_tentative).value
            inv_prepared = ws.cell(row=r_idx, column=col_inv_prepared).value
            inv_sent = ws.cell(row=r_idx, column=col_inv_sent).value
            inv_time = ws.cell(row=r_idx, column=col_inv_time).value
            remarks = ws.cell(row=r_idx, column=col_remarks).value if col_remarks else None
            
            # --- Attendance Tracker Record ---
            invoice_records.append({
                'site_name': site_name,
                'company_name': comp_name,
                'billing_cycle': cycle,
                'ops_incharge': ops_incharge,
                'hr_incharge': hr_incharge,
                'invoice_incharge': invoice_incharge,
                'manager_tentative_date': sql_date(mgr_tentative),
                'manager_received_date': sql_date(mgr_received),
                'hr_tentative_date': sql_date(hr_tentative),
                'hr_received_date': sql_date(hr_received),
                'attendance_received_time': sql_time(att_time),
                'invoice_sharing_tentative_date': sql_date(inv_sharing),
                'invoice_prepared_date': sql_date(inv_prepared),
                'invoice_sent_date': sql_date(inv_sent),
                'invoice_sent_time': sql_time(inv_time),
                'invoice_sent_method_remarks': sql_str(clean_str(remarks)),
                'created_by': creator_user['id'],
                'created_by_name': creator_user['name'],
                'created_by_role': creator_user['role'],
                'created_at': created_at_val,
                'sheet': sname
            })
            
            # --- Finance Tracker Record ---
            gross_val = ws.cell(row=r_idx, column=col_gross).value if col_gross else 0
            admin_chg_val = ws.cell(row=r_idx, column=col_admin_chg).value if col_admin_chg else 0
            cost_contract_val = ws.cell(row=r_idx, column=col_cost_contract).value if col_cost_contract else 0
            cost_admin_val = ws.cell(row=r_idx, column=col_cost_admin_fee).value if col_cost_admin_fee else 0
            admin_remarks_val = ws.cell(row=r_idx, column=col_admin_remarks).value if col_admin_remarks else ''
            
            finance_records.append({
                'site_name': site_name,
                'company_name': comp_name,
                'billing_month': billing_month,
                'contract_amount': sql_num(cost_contract_val),
                'contract_management_fee': sql_num(cost_admin_val),
                'billed_amount': sql_num(gross_val),
                'billed_management_fee': sql_num(admin_chg_val),
                'remarks': sql_str(clean_str(admin_remarks_val)),
                'created_by': creator_user['id'],
                'created_by_name': creator_user['name'],
                'created_by_role': creator_user['role'],
                'created_at': created_at_val,
                'sheet': sname
            })

    wb.close()
    
    print(f"Total site_invoice_tracker records parsed: {len(invoice_records)}")
    print(f"Total site_finance_tracker records parsed: {len(finance_records)}")

    # Generate SQL
    lines = []
    lines.append("-- ==========================================================================")
    lines.append("-- MIGRATION DATA SQL: Invoice Monthly Update Report (Jan '26 to June '26)")
    lines.append(f"-- Source: {EXCEL_PATH}")
    lines.append(f"-- Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- Attendance Tracker records: {len(invoice_records)}")
    lines.append(f"-- Monthly Finance Tracker records: {len(finance_records)}")
    lines.append("-- ==========================================================================\n")
    lines.append("BEGIN;\n")

    # 1. Migrate site_invoice_tracker
    lines.append("-- --------------------------------------------------------------------------")
    lines.append("-- 1. Insert/Upsert into public.site_invoice_tracker (Attendance Tracker Tab)")
    lines.append("-- --------------------------------------------------------------------------\n")
    
    lines.append("""INSERT INTO public.site_invoice_tracker (
    site_name, company_name, billing_cycle,
    ops_incharge, hr_incharge, invoice_incharge,
    manager_tentative_date, manager_received_date,
    hr_tentative_date, hr_received_date, attendance_received_time,
    invoice_sharing_tentative_date, invoice_prepared_date,
    invoice_sent_date, invoice_sent_time, invoice_sent_method_remarks,
    created_by, created_by_name, created_by_role, created_at, updated_at
)
SELECT 
    v.site_name, v.company_name, v.billing_cycle,
    v.ops_incharge, v.hr_incharge, v.invoice_incharge,
    v.manager_tentative_date, v.manager_received_date,
    v.hr_tentative_date, v.hr_received_date, v.attendance_received_time,
    v.invoice_sharing_tentative_date, v.invoice_prepared_date,
    v.invoice_sent_date, v.invoice_sent_time, v.invoice_sent_method_remarks,
    v.created_by, v.created_by_name, v.created_by_role, v.created_at, v.created_at
FROM (
  VALUES""")

    # Format values for site_invoice_tracker with explicit casts on first row
    inv_vals = []
    for idx, r in enumerate(invoice_records):
        if idx == 0:
            row_str = f"    ({sql_str(r['site_name'])}::text, {sql_str(r['company_name'])}::text, {sql_str(r['billing_cycle'])}::text, {sql_str(r['ops_incharge'])}::text, {sql_str(r['hr_incharge'])}::text, {sql_str(r['invoice_incharge'])}::text, {r['manager_tentative_date']}::date, {r['manager_received_date']}::date, {r['hr_tentative_date']}::date, {r['hr_received_date']}::date, {r['attendance_received_time']}::text, {r['invoice_sharing_tentative_date']}::date, {r['invoice_prepared_date']}::date, {r['invoice_sent_date']}::date, {r['invoice_sent_time']}::text, {r['invoice_sent_method_remarks']}::text, '{r['created_by']}'::uuid, '{r['created_by_name']}'::text, '{r['created_by_role']}'::text, '{r['created_at']}'::timestamptz)"
        else:
            row_str = f"    ({sql_str(r['site_name'])}, {sql_str(r['company_name'])}, {sql_str(r['billing_cycle'])}, {sql_str(r['ops_incharge'])}, {sql_str(r['hr_incharge'])}, {sql_str(r['invoice_incharge'])}, {r['manager_tentative_date']}, {r['manager_received_date']}, {r['hr_tentative_date']}, {r['hr_received_date']}, {r['attendance_received_time']}, {r['invoice_sharing_tentative_date']}, {r['invoice_prepared_date']}, {r['invoice_sent_date']}, {r['invoice_sent_time']}, {r['invoice_sent_method_remarks']}, '{r['created_by']}', '{r['created_by_name']}', '{r['created_by_role']}', '{r['created_at']}')"
        inv_vals.append(row_str)

    lines.append(",\n".join(inv_vals))
    lines.append(""") AS v(
    site_name, company_name, billing_cycle,
    ops_incharge, hr_incharge, invoice_incharge,
    manager_tentative_date, manager_received_date,
    hr_tentative_date, hr_received_date, attendance_received_time,
    invoice_sharing_tentative_date, invoice_prepared_date,
    invoice_sent_date, invoice_sent_time, invoice_sent_method_remarks,
    created_by, created_by_name, created_by_role, created_at
)
WHERE NOT EXISTS (
    SELECT 1 FROM public.site_invoice_tracker sit
    WHERE sit.site_name = v.site_name
      AND sit.manager_tentative_date = v.manager_tentative_date
      AND sit.deleted_at IS NULL
);\n""")

    # 2. Migrate site_finance_tracker
    lines.append("-- --------------------------------------------------------------------------")
    lines.append("-- 2. Insert/Upsert into public.site_finance_tracker (Monthly Invoice Tracker Tab)")
    lines.append("-- --------------------------------------------------------------------------\n")
    
    lines.append("""INSERT INTO public.site_finance_tracker (
    site_name, company_name, billing_month,
    contract_amount, contract_management_fee,
    billed_amount, billed_management_fee,
    remarks, status,
    created_by, created_by_name, created_by_role, created_at, updated_at
)
SELECT 
    v.site_name, v.company_name, v.billing_month,
    v.contract_amount, v.contract_management_fee,
    v.billed_amount, v.billed_management_fee,
    v.remarks, 'pending',
    v.created_by, v.created_by_name, v.created_by_role, v.created_at, v.created_at
FROM (
  VALUES""")

    fin_vals = []
    for idx, r in enumerate(finance_records):
        if idx == 0:
            row_str = f"    ({sql_str(r['site_name'])}::text, {sql_str(r['company_name'])}::text, '{r['billing_month']}'::date, {r['contract_amount']}::numeric, {r['contract_management_fee']}::numeric, {r['billed_amount']}::numeric, {r['billed_management_fee']}::numeric, {r['remarks']}::text, '{r['created_by']}'::uuid, '{r['created_by_name']}'::text, '{r['created_by_role']}'::text, '{r['created_at']}'::timestamptz)"
        else:
            row_str = f"    ({sql_str(r['site_name'])}, {sql_str(r['company_name'])}, '{r['billing_month']}', {r['contract_amount']}, {r['contract_management_fee']}, {r['billed_amount']}, {r['billed_management_fee']}, {r['remarks']}, '{r['created_by']}', '{r['created_by_name']}', '{r['created_by_role']}', '{r['created_at']}')"
        fin_vals.append(row_str)

    lines.append(",\n".join(fin_vals))
    lines.append(""") AS v(
    site_name, company_name, billing_month,
    contract_amount, contract_management_fee,
    billed_amount, billed_management_fee,
    remarks, created_by, created_by_name, created_by_role, created_at
)
WHERE NOT EXISTS (
    SELECT 1 FROM public.site_finance_tracker sft
    WHERE sft.site_name = v.site_name
      AND sft.billing_month = v.billing_month
      AND sft.deleted_at IS NULL
);\n""")

    lines.append("COMMIT;\n")
    
    # Save SQL file
    with open(OUTPUT_SQL_PATH, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))
        
    print(f"Successfully generated migration SQL: {OUTPUT_SQL_PATH}")
    print(f"Total SQL lines: {len(lines)}")

if __name__ == '__main__':
    process()
